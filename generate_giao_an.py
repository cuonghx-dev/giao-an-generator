#!/usr/bin/env python3
"""
Giáo Án Generator - Generates lesson plan Excel files.

Combines a schedule file (.xlsx) with content files (.zip of .xlsx)
to produce a lesson plan with 5 sheets (Mon-Fri).

See docs/workflow.md for full documentation.
"""

import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os
import sys
import re
import io
import zipfile
import tempfile
import unicodedata
from datetime import datetime, timedelta

# Map schedule activity types to content files.
# Keys = activity type names as they appear in the schedule.
# Values = corresponding content file in CONTENT_DIR.
ACTIVITY_FILE_MAP = {
    'Trò chuyện đầu ngày': 'Trò chuyện đầu ngày.xlsx',
    'Hoạt động chính': 'Hoạt động chính.xlsx',
    'Hoạt động Tiếng Việt': 'Hoạt động Tiếng Việt.xlsx',
    'Hoạt động tập thể': 'Hoạt động tập thể.xlsx',
    'Hoạt động thể chất': 'Hoạt động tập thể.xlsx',
    'Hoạt động ngoài trời': 'Hoạt động ngoài trời.xlsx',
    'Hoạt động vui chơi': 'Hoạt động vui chơi.xlsx',
    'Chơi theo chủ đề': 'Hoạt động vui chơi.xlsx',
    'Hoạt động Dự án': 'Hoạt động dự án.xlsx',
    'Hoạt động dự án': 'Hoạt động dự án.xlsx',
    'Dã ngoại/Sự kiện gia đình': 'Dã ngoại sự kiện và gia đình.xlsx',
}

# Day columns in schedule file (weekday_index -> column_number).
# Mon-Fri maps to columns C-G in the schedule sheet.
DAY_COLUMNS = {0: 3, 1: 4, 2: 5, 3: 6, 4: 7}  # Mon=C, Tue=D, Wed=E, Thu=F, Fri=G

# Time slots in schedule: (type_row, detail_row).
# Each pair = one time slot. type_row has the activity category,
# detail_row has the specific activity name.
TIME_SLOTS = [(8, 9), (10, 11), (12, 13), (14, 15), (21, 22), (23, 24)]

# Keywords to skip in schedule — if any keyword appears in the
# activity type or detail, the activity is excluded from the lesson plan.
SKIP_KEYWORDS = [
    'Story time', 'ESL Class', 'ESL 1', 'ESL 2', 'ESL 3',
    'Chơi tự do', 'Vui chơi tự do', 'Vui chơi Softplay', 'Softplay',
    'phòng Lego', 'phòng Cosplay', 'phòng thư viện',
    'Đón trẻ', 'Thể dục sáng', 'Ăn sáng', 'Ăn trưa',
    'Ngủ', 'Vệ sinh', 'Ăn chiều', 'Uống sữa',
    'Tái hiện', 'Chơi tự chọn', 'Dọn dẹp',
    'Hoạt động thư viện',
]

# ============================================================
# STYLES
# ============================================================
MEDIUM_SIDE = Side(style='medium')
THIN_SIDE = Side(style='thin')
BORDER_MEDIUM = Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE, bottom=MEDIUM_SIDE)
BORDER_THIN = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

FONT_TITLE = Font(name='Cambria', size=28, color='FF000000')
FONT_NORMAL = Font(name='Cambria', size=10, color='FF000000')
FONT_BOLD = Font(name='Cambria', size=10, bold=True, color='FF000000')
FONT_GRAY_BOLD = Font(name='Cambria', size=10, bold=True, color='FF5A5A5A')
FONT_WHITE_BOLD = Font(name='Cambria', size=10, bold=True, color='FFFFFFFF')
FONT_FOOTER = Font(name='Cambria', size=11, bold=True)

FILL_RED = PatternFill(start_color='FFCC0000', end_color='FFCC0000', fill_type='solid')
FILL_GREEN = PatternFill(start_color='FFB6D7A8', end_color='FFB6D7A8', fill_type='solid')
FILL_NAVY = PatternFill(start_color='FF1F3864', end_color='FF1F3864', fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(vertical='center', wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)


# ============================================================
# CONTENT FILE PARSER
# ============================================================

def nfc(text):
    """Normalize Unicode to NFC form for consistent comparison."""
    return unicodedata.normalize('NFC', str(text))


def normalize_name(name):
    """Normalize activity name for fuzzy matching."""
    name = nfc(str(name)).strip().strip('"\'""\u201c\u201d')
    # Remove common prefixes
    prefixes = [
        'Hoạt động ngoài trời: ', 'Hoạt động ngoài trời:',
        'Hoạt động: ', 'Truyện: ', 'Truyện ', 'Thơ: ', 'Thơ ',
        'Câu chuyện của tháng\n', 'Kỹ năng ngôn ngữ\n', 'Kỹ năng ngôn ngữ \n',
        'Sức khoẻ thể chất\n', 'Kỹ năng toán học\n',
        'Cảm xúc xã hội\n', 'Giáo dục Khoa học và môi trường\n',
    ]
    for prefix in prefixes:
        if name.startswith(nfc(prefix)):
            name = name[len(nfc(prefix)):]
    name = name.strip().strip('"\'""\u201c\u201d')
    return name


def match_score(a, b):
    """Compute similarity between two normalized names. Higher = better match."""
    a, b = nfc(a).lower(), nfc(b).lower()
    if a == b:
        return 100
    if a in b or b in a:
        return 80 + min(len(a), len(b))
    # Check word overlap
    words_a = set(a.split())
    words_b = set(b.split())
    if not words_a or not words_b:
        return 0
    overlap = words_a & words_b
    return len(overlap) * 20


def parse_all_content_files(content_path):
    """Parse all content files from a .zip or directory. Returns dict of name -> activity block."""
    all_activities = {}

    if zipfile.is_zipfile(content_path):
        _parse_from_zip(content_path, all_activities)
    elif os.path.isdir(content_path):
        _parse_from_dir(content_path, all_activities)
    else:
        print(f"Lỗi: '{content_path}' không phải file zip hoặc thư mục hợp lệ.")
        sys.exit(1)

    return all_activities


def _parse_from_zip(zip_path, all_activities):
    """Extract and parse content files from a .zip archive."""
    with zipfile.ZipFile(zip_path, 'r') as zf:
        for entry in zf.namelist():
            # Skip Mac OS resource forks and non-xlsx files
            if '__MACOSX' in entry or not entry.endswith('.xlsx'):
                continue
            xlsx_bytes = zf.read(entry)
            # Load with data_only=True
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            ws = wb.active
            # Also load with formulas for fallback
            wb_formulas = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
            ws_formulas = wb_formulas.active

            blocks = find_activity_blocks(ws, ws_formulas)
            for block in blocks:
                name = nfc(block['name'])
                block['name'] = name
                norm = normalize_name(name)
                all_activities[name] = block
                all_activities[norm] = block


def _parse_from_dir(dir_path, all_activities):
    """Parse content files from a directory of .xlsx files."""
    for fname in os.listdir(dir_path):
        if not fname.endswith('.xlsx'):
            continue
        fpath = os.path.join(dir_path, fname)
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        wb_formulas = openpyxl.load_workbook(fpath)
        ws_formulas = wb_formulas.active

        blocks = find_activity_blocks(ws, ws_formulas)
        for block in blocks:
            name = nfc(block['name'])
            block['name'] = name
            norm = normalize_name(name)
            all_activities[name] = block
            all_activities[norm] = block


def find_activity_blocks(ws, ws_formulas):
    """Find activity blocks using column A merged cells."""
    blocks = []
    skip_values = {'HOẠT ĐỘNG', 'STT', 'GIÁO ÁN', 'Suy ngẫm về dạy & học:'}

    # Collect column A merged ranges
    merged_a = sorted(
        [m for m in ws.merged_cells.ranges if m.min_col == 1 and m.max_col == 1],
        key=lambda m: m.min_row
    )

    for merge in merged_a:
        val = ws.cell(row=merge.min_row, column=1).value
        if not val or isinstance(val, (int, float)):
            continue
        val_str = str(val).strip()
        if val_str in skip_values:
            continue

        block = extract_block(ws, ws_formulas, merge.min_row, merge.max_row)
        block['name'] = val_str
        blocks.append(block)

    return blocks


def extract_block(ws, ws_formulas, start_row, end_row):
    """Extract activity block data, values, and merge patterns."""
    rows_data = []
    for row in range(start_row, end_row + 1):
        row_data = {}
        for col in range(1, 10):
            val = ws.cell(row=row, column=col).value
            # Fallback to formula sheet if data_only gave None but formula exists
            if val is None:
                formula_val = ws_formulas.cell(row=row, column=col).value
                if formula_val and isinstance(formula_val, str) and formula_val.startswith('='):
                    # Try to resolve simple cell references like =H10
                    ref_match = re.match(r'^=([A-Z]+)(\d+)$', formula_val)
                    if ref_match:
                        ref_col = openpyxl.utils.column_index_from_string(ref_match.group(1))
                        ref_row = int(ref_match.group(2))
                        val = ws.cell(row=ref_row, column=ref_col).value
            row_data[col] = val
        rows_data.append(row_data)

    # Extract merge patterns relative to block start
    merges = []
    for m in ws.merged_cells.ranges:
        if m.min_row >= start_row and m.max_row <= end_row and m.min_col <= 9:
            merges.append({
                'rel_min_row': m.min_row - start_row,
                'rel_max_row': m.max_row - start_row,
                'min_col': m.min_col,
                'max_col': m.max_col,
            })

    # Collect all column C values for summary objectives
    c_values = []
    for rd in rows_data:
        if rd[3]:  # col C
            c_values.append(str(rd[3]).strip())

    return {
        'num_rows': end_row - start_row + 1,
        'rows': rows_data,
        'merges': merges,
        'objectives_summary': '\n'.join(c_values),
    }


# ============================================================
# SCHEDULE PARSER
# ============================================================

def parse_schedule(schedule_path, class_name):
    """Parse weekly schedule for a given class."""
    wb = openpyxl.load_workbook(schedule_path)

    # Find sheet by name (handle trailing spaces)
    sheet_name = None
    for s in wb.sheetnames:
        if s.strip() == class_name.strip():
            sheet_name = s
            break
    if not sheet_name:
        available = ', '.join(s.strip() for s in wb.sheetnames)
        print(f"Lớp '{class_name}' không tìm thấy. Các lớp có sẵn: {available}")
        sys.exit(1)

    ws = wb[sheet_name]

    # Parse week info (row 2)
    week_info = str(ws.cell(row=2, column=2).value or '')
    date_match = re.search(r'(\d+)/(\d+)\s*đến\s*(\d+)/(\d+)/(\d+)', week_info)
    if not date_match:
        print(f"Không thể phân tích ngày từ: {week_info}")
        sys.exit(1)
    start_day = int(date_match.group(1))
    start_month = int(date_match.group(2))
    year = int(date_match.group(5))
    start_date = datetime(year, start_month, start_day)

    # Parse theme (row 3) - extract Vietnamese name only, remove "(Tuần X)" etc.
    theme_raw = str(ws.cell(row=3, column=2).value or '')
    theme_match = re.match(r'Chủ đề/Theme:\s*(.+?)(?:\s*\(|/|$)', theme_raw)
    theme = theme_match.group(1).strip() if theme_match else theme_raw.strip()

    # Parse teacher (row 3, col G) - get first teacher name
    teacher_raw = str(ws.cell(row=3, column=7).value or '')
    teacher_match = re.search(r'Giáo viên:\s*(.+)', teacher_raw)
    teacher_list = teacher_match.group(1).strip().rstrip("' ") if teacher_match else ''
    teacher = teacher_list.split('-')[0].strip() if teacher_list else ''

    # Parse activities per day
    days = []
    for day_idx in range(5):
        col = DAY_COLUMNS[day_idx]
        date = start_date + timedelta(days=day_idx)
        day_num = day_idx + 2  # Thứ 2 = Monday

        activities = []
        for type_row, detail_row in TIME_SLOTS:
            type_val = ws.cell(row=type_row, column=col).value
            detail_val = ws.cell(row=detail_row, column=col).value

            if not type_val:
                continue

            type_str = str(type_val).strip()
            detail_str = str(detail_val).strip() if detail_val else ''

            if should_skip(type_str, detail_str):
                continue

            # Clean up activity type
            type_clean = re.sub(r'\s*\(.*?\)\s*', '', type_str).strip()

            activities.append({
                'type': type_clean,
                'detail': detail_str,
            })

        days.append({
            'day_num': day_num,
            'date': date,
            'activities': activities,
        })

    return {
        'class_name': class_name,
        'theme': theme,
        'teacher': teacher,
        'days': days,
    }


def should_skip(type_str, detail_str):
    """Check if a schedule activity should be skipped."""
    combined = nfc(type_str + ' ' + detail_str)
    for kw in SKIP_KEYWORDS:
        if nfc(kw) in combined:
            return True

    # Must map to a known activity type
    type_clean = re.sub(r'\s*\(.*?\)\s*', '', type_str).strip()
    # Try fuzzy type matching
    if type_clean not in ACTIVITY_FILE_MAP:
        # Try matching with NFC normalization
        found = False
        for key in ACTIVITY_FILE_MAP:
            if nfc(type_clean) == nfc(key):
                found = True
                break
        if not found:
            return True

    return False


# ============================================================
# ACTIVITY MATCHING
# ============================================================

def find_content(activity, all_content):
    """Find matching content block using 4-level fuzzy matching."""
    detail = nfc(activity['detail'])

    # Build list of candidate names to try (from schedule detail)
    candidates = [detail, normalize_name(detail)]
    if '\n' in detail:
        for part in detail.split('\n'):
            part = part.strip()
            if part:
                candidates.extend([part, normalize_name(part)])

    # Try exact match (NFC-normalized)
    for candidate in candidates:
        for stored_name, block in all_content.items():
            if nfc(candidate) == nfc(stored_name):
                return block

    # Try normalized match
    for candidate in candidates:
        norm_cand = normalize_name(candidate)
        for stored_name, block in all_content.items():
            norm_stored = normalize_name(stored_name)
            if nfc(norm_cand) == nfc(norm_stored):
                return block

    # Fuzzy substring match
    for candidate in candidates:
        norm_cand = nfc(normalize_name(candidate)).lower()
        if len(norm_cand) < 4:
            continue
        for stored_name, block in all_content.items():
            norm_stored = nfc(normalize_name(stored_name)).lower()
            if len(norm_stored) < 4:
                continue
            if norm_cand in norm_stored or norm_stored in norm_cand:
                return block

    # Word-overlap scoring as last resort
    best_score = 0
    best_block = None
    for candidate in candidates:
        norm_cand = normalize_name(candidate)
        for stored_name, block in all_content.items():
            norm_stored = normalize_name(stored_name)
            score = match_score(norm_cand, norm_stored)
            if score > best_score and score >= 60:
                best_score = score
                best_block = block

    return best_block


# ============================================================
# SHEET BUILDER
# ============================================================

def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    """Apply styles to a cell."""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def style_range(ws, min_row, max_row, min_col, max_col, font=None, fill=None, alignment=None, border=None):
    """Apply styles to a rectangular range of cells."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            apply_style(ws.cell(row=row, column=col), font, fill, alignment, border)


def write_header(ws, schedule, day, start_row=1):
    """Write header section (rows 1-7). Returns next row."""
    r = start_row
    date = day['date']
    day_num = day['day_num']

    # Row 1: Title
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value='GIÁO ÁN')
    apply_style(c, FONT_TITLE, alignment=ALIGN_CENTER, border=BORDER_MEDIUM)
    style_range(ws, r, r, 2, 9, border=BORDER_MEDIUM)
    ws.row_dimensions[r].height = 48
    r += 1

    # Row 2: Date
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    date_str = f'Dành cho thứ {day_num}, ngày {date.strftime("%d/%m/")} {date.year}'
    c = ws.cell(row=r, column=1, value=date_str)
    apply_style(c, FONT_GRAY_BOLD, alignment=ALIGN_CENTER, border=BORDER_MEDIUM)
    style_range(ws, r, r, 2, 9, border=BORDER_MEDIUM)
    r += 1

    # Row 3: Theme
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value=f'Bài 1 – Chương: {schedule["theme"]}')
    apply_style(c, Font(name='Cambria', size=10, color='FFFFFFFF'), FILL_RED, ALIGN_CENTER, BORDER_MEDIUM)
    style_range(ws, r, r, 2, 9, border=BORDER_MEDIUM)
    r += 1

    # Row 4: Location
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value='Cơ sở: Times City ')
    apply_style(c, FONT_BOLD, alignment=ALIGN_LEFT_CENTER, border=BORDER_MEDIUM)
    style_range(ws, r, r, 2, 9, border=BORDER_MEDIUM)
    r += 1

    # Row 5: Teacher
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value=f'Giáo viên thực hiện: {schedule["teacher"]}')
    apply_style(c, FONT_BOLD, alignment=ALIGN_LEFT_CENTER, border=BORDER_MEDIUM)
    style_range(ws, r, r, 2, 9, border=BORDER_MEDIUM)
    r += 1

    # Row 6: Class
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value=f'Lớp: {schedule["class_name"]}')
    apply_style(c, FONT_BOLD, alignment=ALIGN_LEFT_CENTER, border=BORDER_MEDIUM)
    style_range(ws, r, r, 2, 9, border=BORDER_MEDIUM)
    r += 1

    # Row 7: Activities header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value='Các hoạt động trong ngày:')
    apply_style(c, FONT_BOLD, FILL_GREEN, ALIGN_CENTER, BORDER_MEDIUM)
    style_range(ws, r, r, 2, 9, border=BORDER_MEDIUM)
    r += 1

    return r


def write_summary_table(ws, matched_activities, start_row):
    """Write summary table. Returns next row."""
    r = start_row

    # Header row
    headers = ['STT', 'Loại hoạt động', 'Tên hoạt động', 'Mục tiêu hoạt động']
    for i, h in enumerate(headers):
        col = i + 1
        if i == 3:  # Merge D:I for objectives header
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
        ws.cell(row=r, column=col, value=h)
        apply_style(ws.cell(row=r, column=col), FONT_BOLD, alignment=ALIGN_CENTER, border=BORDER_MEDIUM)
    style_range(ws, r, r, 5, 9, border=BORDER_MEDIUM)
    r += 1

    # Activity rows
    for idx, act in enumerate(matched_activities):
        ws.cell(row=r, column=1, value=idx + 1)
        apply_style(ws.cell(row=r, column=1), FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_MEDIUM)

        ws.cell(row=r, column=2, value=act['type'])
        apply_style(ws.cell(row=r, column=2), FONT_NORMAL, alignment=ALIGN_LEFT_CENTER, border=BORDER_MEDIUM)

        # Clean detail for summary display
        detail_display = act['detail']
        # For TC&SK entries, extract just the activity name
        if 'TC&SK' in detail_display or 'Vivokids' in detail_display:
            for line in detail_display.split('\n'):
                line = line.strip()
                if line.startswith('TC& SK:') or line.startswith('TC&SK:'):
                    detail_display = line.split(':', 1)[1].strip()
                    break

        ws.cell(row=r, column=3, value=detail_display)
        apply_style(ws.cell(row=r, column=3), FONT_NORMAL, alignment=ALIGN_LEFT_CENTER, border=BORDER_MEDIUM)

        # Objectives from content or empty
        objectives = ''
        if act.get('content'):
            objectives = act['content']['objectives_summary']
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
        ws.cell(row=r, column=4, value=objectives)
        apply_style(ws.cell(row=r, column=4), FONT_NORMAL, alignment=ALIGN_LEFT_CENTER, border=BORDER_MEDIUM)
        style_range(ws, r, r, 5, 9, border=BORDER_MEDIUM)
        r += 1

    return r


def write_detail_header(ws, start_row):
    """Write detail table header (2 rows). Returns next row."""
    r = start_row

    # Row 1 of header
    headers_r1 = {
        1: 'HOẠT ĐỘNG', 2: 'CHUẨN/MỤC TIÊU CHƯƠNG',
        3: 'MỤC TIÊU BÀI/ MỤC TIÊU HOẠT ĐỘNG', 4: 'TIÊU CHÍ THÀNH CÔNG',
        5: 'TÀI LIỆU HỌC TẬP, CHUẨN BỊ', 6: 'CÁC BƯỚC TỔ CHỨC HOẠT ĐỘNG',
        8: 'ĐÁNH GIÁ', 9: 'PHÂN HÓA',
    }

    # Merge vertically for most columns (rows r and r+1)
    for col in [1, 2, 3, 4, 5, 8, 9]:
        ws.merge_cells(start_row=r, start_column=col, end_row=r + 1, end_column=col)
    # Merge F:G horizontally on row 1
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

    for col, text in headers_r1.items():
        ws.cell(row=r, column=col, value=text)
        apply_style(ws.cell(row=r, column=col), FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER, BORDER_MEDIUM)

    # Apply style to col 7 (merged with 6)
    apply_style(ws.cell(row=r, column=7), FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER, BORDER_MEDIUM)
    r += 1

    # Row 2 of header: sub-headers for F and G
    ws.cell(row=r, column=6, value='Hoạt động')
    apply_style(ws.cell(row=r, column=6), FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER,
                Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE))

    ws.cell(row=r, column=7, value='Mục tiêu hướng tới')
    apply_style(ws.cell(row=r, column=7), FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER,
                Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE))

    # Style remaining cells in row 2
    for col in [1, 2, 3, 4, 5, 8, 9]:
        apply_style(ws.cell(row=r, column=col), FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER, BORDER_MEDIUM)
    r += 1

    return r


def write_activity_block(ws, activity, start_row):
    """Write a single activity detail block. Returns next row."""
    content = activity.get('content')

    if content:
        return write_content_block(ws, content, start_row)
    else:
        return write_placeholder_block(ws, activity, start_row)


def write_content_block(ws, content, target_start):
    """Write activity content block from content file data."""
    num_rows = content['num_rows']

    # Write cell values
    for i, row_data in enumerate(content['rows']):
        target_row = target_start + i
        for col in range(1, 10):
            val = row_data.get(col)
            if val is not None:
                ws.cell(row=target_row, column=col, value=val)
            apply_style(ws.cell(row=target_row, column=col),
                        FONT_NORMAL, alignment=ALIGN_LEFT_TOP, border=BORDER_THIN)

    # Apply merge patterns
    for m in content['merges']:
        try:
            ws.merge_cells(
                start_row=target_start + m['rel_min_row'],
                start_column=m['min_col'],
                end_row=target_start + m['rel_max_row'],
                end_column=m['max_col'],
            )
        except Exception:
            pass  # Skip if merge conflicts

    return target_start + num_rows


def write_placeholder_block(ws, activity, start_row):
    """Write a 1-row placeholder for activities without content."""
    r = start_row

    # Activity name: type + cleaned detail
    detail = activity.get('detail', '')
    act_type = activity.get('type', '')
    # Extract meaningful name from detail (last part after \n usually)
    detail_parts = [p.strip() for p in detail.split('\n') if p.strip()]
    clean_name = detail_parts[-1] if detail_parts else detail
    name = f"{act_type}\n{clean_name}" if act_type else clean_name

    ws.cell(row=r, column=1, value=name)
    for col in range(1, 10):
        apply_style(ws.cell(row=r, column=col), FONT_NORMAL, alignment=ALIGN_LEFT_TOP, border=BORDER_THIN)

    return r + 1


def write_footer(ws, start_row):
    """Write footer section."""
    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=8)
    ws.cell(row=r, column=1, value='Suy ngẫm về dạy & học:')
    apply_style(ws.cell(row=r, column=1), FONT_FOOTER, alignment=Alignment(wrap_text=True), border=BORDER_MEDIUM)

    # Style remaining cells in footer
    for row in range(r, r + 2):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, font=FONT_FOOTER, border=BORDER_MEDIUM)
            cell.alignment = Alignment(wrap_text=True)

    return r + 2


def set_sheet_properties(ws):
    """Set column widths, page setup, and sheet format to match template."""
    # Column widths
    ws.column_dimensions['A'].width = 25.71
    ws.column_dimensions['J'].width = 8.71
    # B-I use default column width (14.43)

    # Sheet format
    ws.sheet_format.defaultColWidth = 14.43
    ws.sheet_format.defaultRowHeight = 15.0

    # Page setup - landscape orientation
    ws.page_setup.orientation = 'landscape'

    # Page margins matching original
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


def build_sheet(ws, schedule, day, matched_activities):
    """Build a complete giao_an sheet for one weekday."""
    r = 1

    # Header (rows 1-7)
    r = write_header(ws, schedule, day, r)

    # Summary table (row 8 header + 1 row per activity)
    r = write_summary_table(ws, matched_activities, r)

    # Detail table header (2 rows: column headers + F/G sub-headers)
    r = write_detail_header(ws, r)

    # Activity detail blocks (variable rows per activity)
    for act in matched_activities:
        r = write_activity_block(ws, act, r)

    # Footer (2 rows: "Suy ngẫm về dạy & học:")
    write_footer(ws, r)

    # Sheet properties (columns, page setup, format)
    set_sheet_properties(ws)

    # Row heights: row 1 = 48, rows 2 through data = 12.75, rest uses default (15.0)
    for row in range(2, r + 3):
        ws.row_dimensions[row].height = 12.75


# ============================================================
# MAIN
# ============================================================

def list_classes(schedule_path):
    """Print available class names from the schedule file."""
    wb = openpyxl.load_workbook(schedule_path)
    print(f"Các lớp có sẵn trong '{schedule_path}':")
    for name in wb.sheetnames:
        print(f"  - {name.strip()}")


def get_all_classes(schedule_path):
    """Return list of class names from the schedule file."""
    wb = openpyxl.load_workbook(schedule_path)
    return [name.strip() for name in wb.sheetnames]


def generate(schedule_path, content_path, class_names, output_dir):
    """Generate lesson plan files for one or more classes."""
    print(f"Đang đọc nội dung giảng dạy từ '{content_path}'...")
    all_content = parse_all_content_files(content_path)
    print(f"  Tìm thấy {len(set(id(v) for v in all_content.values()))} hoạt động")

    os.makedirs(output_dir, exist_ok=True)

    for class_name in class_names:
        output_path = os.path.join(output_dir, f"giao_an_{class_name.replace(' ', '_')}.xlsx")
        print(f"\n{'='*60}")
        print(f"Đang đọc kế hoạch giảng dạy cho lớp '{class_name}'...")
        schedule = parse_schedule(schedule_path, class_name)

        print(f"Đang tạo giáo án...")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for day in schedule['days']:
            day_num = day['day_num']
            date = day['date']
            sheet_name = f"Thứ {day_num} ngày {date.day}{date.month}"

            matched = []
            for act in day['activities']:
                content = find_content(act, all_content)
                matched.append({**act, 'content': content})

                status = '✓' if content else '✗ (không có nội dung)'
                print(f"  {sheet_name}: {act['type']} - {act['detail'][:40]}... {status}")

            if not matched:
                print(f"  {sheet_name}: Không có hoạt động nào")
                continue

            ws = wb.create_sheet(title=sheet_name)
            build_sheet(ws, schedule, day, matched)

        wb.save(output_path)
        print(f"Đã tạo: {output_path}")

    print(f"\nHoàn thành! Đã tạo {len(class_names)} file trong '{output_dir}/'.")


def main():
    parser = argparse.ArgumentParser(
        description='Giáo Án Generator - Tạo giáo án từ kế hoạch giảng dạy và nội dung bài học.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ví dụ:
  %(prog)s -s ke_hoach.xlsx -c noi_dung.zip --class "Vega 12"
  %(prog)s -s ke_hoach.xlsx -c noi_dung.zip --all
  %(prog)s -s ke_hoach.xlsx -c noi_dung.zip --all -o output/
  %(prog)s -s ke_hoach.xlsx --list-classes

Xem docs/workflow.md để biết thêm chi tiết.
        """,
    )
    parser.add_argument(
        '--schedule', '-s',
        required=True,
        metavar='FILE',
        help='Đường dẫn file kế hoạch giảng dạy (.xlsx)',
    )
    parser.add_argument(
        '--content', '-c',
        metavar='PATH',
        help='Đường dẫn nội dung bài học (.zip hoặc thư mục chứa các file .xlsx)',
    )

    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        '--class', '-C',
        dest='class_name',
        metavar='NAME',
        help='Tên lớp (ví dụ: "Vega 12")',
    )
    group.add_argument(
        '--all', '-a',
        action='store_true',
        help='Tạo giáo án cho tất cả các lớp',
    )
    group.add_argument(
        '--list-classes', '-l',
        action='store_true',
        help='Liệt kê các lớp có sẵn trong file kế hoạch giảng dạy',
    )

    parser.add_argument(
        '--output', '-o',
        default='output',
        metavar='DIR',
        help='Thư mục đầu ra (mặc định: output/)',
    )

    args = parser.parse_args()

    # Validate schedule file exists
    if not os.path.isfile(args.schedule):
        parser.error(f"Không tìm thấy file: '{args.schedule}'")

    # List classes mode
    if args.list_classes:
        list_classes(args.schedule)
        return

    # No action specified — show classes as hint
    if not args.content and not args.class_name and not args.all:
        list_classes(args.schedule)
        print("\nSử dụng --class hoặc --all cùng với --content để tạo giáo án.")
        return

    # Validate content arg
    if not args.content:
        parser.error("Cần cung cấp --content (-c) khi tạo giáo án")
    if not os.path.exists(args.content):
        parser.error(f"Không tìm thấy: '{args.content}'")

    # Determine which classes to generate
    if args.all:
        class_names = get_all_classes(args.schedule)
    else:
        class_names = [args.class_name]

    generate(args.schedule, args.content, class_names, args.output)


if __name__ == '__main__':
    main()

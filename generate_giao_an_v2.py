#!/usr/bin/env python3
"""
Giáo Án Generator - Generates lesson plan Excel files.

Combines a weekly schedule file (one sheet per class) with a single
"KH chi tiết" workbook that holds the detailed content of every activity
in a wide block layout, and produces a lesson plan with one sheet per
weekday (Mon-Fri).

The content workbook is laid out like this:

    TUẦN 1                                    <- week header (merged A:X)
    HOẠT ĐỘNG CHÍNH                           <- category header (merged A:X)
    HOẠT ĐỘNG 1: ... | Hoạt động 2: ... | ... <- 5 activity blocks side by side
    GV SOẠN | Tiêu chuẩn | Mục tiêu bài | ... <- field labels per block
    <values>                                  <- one data row per block

Each block spans ~6 columns (A:F, G:L, M:R, S:Y, Z:AE). The "Hoạt động"
field of a block holds the whole lesson: numbered phases plus embedded
"Đánh giá" and "Phân hóa" sections, which are split back out into the
ĐÁNH GIÁ / PHÂN HÓA columns of the output.

See docs/workflow.md for full documentation.
"""

import argparse
import os
import re
import sys
import unicodedata
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Day columns in the schedule sheet: Mon-Fri = C-G.
DAY_COLUMNS = {0: 3, 1: 4, 2: 5, 3: 6, 4: 7}

# Teaching slots in the schedule, as (type_row, detail_row) pairs. The type
# row holds the activity category, the detail row its specific name.
TIME_SLOTS = [(8, 9), (10, 11), (12, 13), (14, 15), (21, 22), (23, 24)]

# A slot is dropped when its type or name contains one of these.
SKIP_KEYWORDS = [
    'Story time', 'ESL Class', 'ESL 1', 'ESL 2', 'ESL 3',
    'Chơi tự do', 'Vui chơi tự do', 'Vui chơi Softplay', 'Softplay',
    'phòng Lego', 'phòng Cosplay', 'phòng thư viện',
    'Đón trẻ', 'Thể dục sáng', 'Ăn sáng', 'Ăn trưa',
    'Ngủ', 'Vệ sinh', 'Ăn chiều', 'Uống sữa',
    'Tái hiện', 'Chơi tự chọn', 'Dọn dẹp',
    'Hoạt động thư viện',
]

# ============================================================
# STYLES
# ============================================================
MEDIUM_SIDE = Side(style='medium')
THIN_SIDE = Side(style='thin')
BORDER_MEDIUM = Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE, bottom=MEDIUM_SIDE)
BORDER_THIN = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

FONT_TITLE = Font(name='Cambria', size=18, bold=True, color='FF000000')
FONT_THEME = Font(name='Cambria', size=10, color='FFFFFFFF')
FONT_NORMAL = Font(name='Cambria', size=10, color='FF000000')
FONT_BOLD = Font(name='Cambria', size=10, bold=True, color='FF000000')
FONT_GRAY_BOLD = Font(name='Cambria', size=10, bold=True, color='FF5A5A5A')
FONT_WHITE_BOLD = Font(name='Cambria', size=10, bold=True, color='FFFFFFFF')
FONT_FOOTER = Font(name='Cambria', size=11, bold=True)

FILL_RED = PatternFill(start_color='FFCC0000', end_color='FFCC0000', fill_type='solid')
FILL_GREEN = PatternFill(start_color='FFB6D7A8', end_color='FFB6D7A8', fill_type='solid')
FILL_NAVY = PatternFill(start_color='FF1F3864', end_color='FF1F3864', fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(vertical='center', wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)

DEFAULT_LOCATION = 'Trường mầm non Vinschool Times City T1 - T2'

# Column widths of the output sheet (col I keeps the default width).
COLUMN_WIDTHS = {
    'A': 19.6, 'B': 25.9, 'C': 35.0, 'D': 26.1,
    'E': 34.0, 'F': 83.1, 'G': 21.4, 'H': 25.9, 'J': 8.9,
}

# Fallback column spans of the 5 activity blocks, used when the activity
# title row has no merged ranges to read the spans from.
DEFAULT_BLOCK_SPANS = [(1, 6), (7, 12), (13, 18), (19, 25), (26, 31)]

# Field label (lowercased) -> key of the parsed activity.
FIELD_LABELS = {
    'gv soạn': 'teacher',
    'giáo viên soạn': 'teacher',
    'giáo viên chuẩn bị': 'prep',
    'tiêu chuẩn': 'standard',
    'chuẩn': 'standard',
    'mục tiêu bài': 'objectives',
    'mục tiêu hoạt động': 'objectives',
    'tiêu chí thành công': 'criteria',
    'hoạt động': 'steps',
    'các bước tổ chức': 'steps',
    'tài liệu học tập': 'materials',
    'đồ dùng': 'materials',
}

# Labels that mark the start of the header row of an activity block.
HEADER_ROW_LABELS = {'gv soạn', 'giáo viên soạn'}

# Schedule activity type -> category header in the content workbook.
TYPE_TO_CATEGORY = {
    'trò chuyện đầu ngày': 'trò chuyện đầu ngày',
    'hoạt động chính': 'hoạt động chính',
    'hoạt động tiếng việt': 'tiếng việt',
    'tiếng việt': 'tiếng việt',
    'chơi theo chủ đề': 'chơi theo chủ đề',
    'hoạt động vui chơi': 'chơi theo chủ đề',
    'hoạt động ngoài trời': 'hoạt động ngoài trời',
    'hoạt động tập thể': 'hoạt động thể chất',
    'hoạt động thể chất': 'hoạt động thể chất',
    'thể chất': 'hoạt động thể chất',
}

# Category header -> label used in column A of the detail table.
CATEGORY_TITLES = {
    'trò chuyện đầu ngày': 'Trò chuyện đầu ngày',
    'hoạt động chính': 'Hoạt động chính',
    'tiếng việt': 'Hoạt động Tiếng Việt',
    'chơi theo chủ đề': 'Chơi theo chủ đề',
    'hoạt động ngoài trời': 'Hoạt động ngoài trời',
    'hoạt động thể chất': 'Hoạt động thể chất',
}

# A numbered line in the "Hoạt động" field starts a new phase (= a new row
# in the output) only when it also mentions one of these. Without this the
# numbered greeting/attendance steps of a morning-chat lesson would each
# become their own row.
PHASE_KEYWORDS = [
    'mở đầu', 'gây hứng thú', 'hứng thú', 'khởi động', 'ổn định',
    'hoạt động chính', 'giảng dạy chính', 'giới thiệu hoạt động',
    'làm mẫu', 'kết thúc', 'củng cố',
]

# Section headings inside the "Hoạt động" field. The heading sometimes runs
# straight into its first sentence ("Đánh giá: Trẻ có tập trung lắng nghe...").
SECTION_HEADINGS = [
    ('assessment', re.compile(r'^\s*đánh giá\b', re.IGNORECASE)),
    ('differentiation', re.compile(r'^\s*phân h(óa|oá)\b', re.IGNORECASE)),
]

NUMBERED_LINE = re.compile(r'^\s*([IVX]+|\d+)\s*[\.\)]\s*(.*)$')
ACTIVITY_TITLE = re.compile(r'^\s*hoạt động\s*\d+\s*[:.]?\s*(.*)$', re.IGNORECASE)
WEEK_HEADER = re.compile(r'^\s*tuần\s*(\d+)\s*$', re.IGNORECASE)


# ============================================================
# TEXT HELPERS
# ============================================================

def nfc(text):
    """Normalize Unicode to NFC form for consistent comparison."""
    return unicodedata.normalize('NFC', str(text))


def normalize_name(name):
    """Normalize an activity name for fuzzy matching."""
    name = nfc(str(name)).strip().strip('"\'\u201c\u201d')
    prefixes = [
        'Hoạt động ngoài trời: ', 'Hoạt động ngoài trời:',
        'Hoạt động: ', 'Truyện: ', 'Truyện ', 'Thơ: ', 'Thơ ',
        'Câu chuyện của tháng\n', 'Kỹ năng ngôn ngữ\n', 'Kỹ năng ngôn ngữ \n',
        'Sức khoẻ thể chất\n', 'Kỹ năng toán học\n',
        'Cảm xúc xã hội\n', 'Giáo dục Khoa học và môi trường\n',
    ]
    for prefix in prefixes:
        if name.startswith(nfc(prefix)):
            name = name[len(nfc(prefix)):]
    return name.strip().strip('"\'\u201c\u201d')


def match_score(a, b):
    """Similarity of two normalized names. Higher = better match."""
    a, b = nfc(a).lower(), nfc(b).lower()
    if a == b:
        return 100
    if a in b or b in a:
        return 80 + min(len(a), len(b))
    words_a, words_b = set(a.split()), set(b.split())
    if not words_a or not words_b:
        return 0
    return len(words_a & words_b) * 20


# ============================================================
# CONTENT PARSER ("KH chi tiết" workbook)
# ============================================================

def label_key(value):
    """Lowercase, whitespace-collapsed form of a cell used for label lookup."""
    return re.sub(r'\s+', ' ', nfc(str(value or '')).strip()).lower()


def is_section_header(ws, row, last_col):
    """A week or category header: a value in column A and nothing else."""
    if ws.cell(row=row, column=1).value is None:
        return False
    for col in range(2, last_col + 1):
        if ws.cell(row=row, column=col).value is not None:
            return False
    return True


def block_spans(ws, title_row, last_col):
    """Column spans of the activity blocks, read from the title row merges."""
    spans = sorted(
        (m.min_col, m.max_col)
        for m in ws.merged_cells.ranges
        if m.min_row == title_row and m.max_row == title_row
    )
    if not spans:
        spans = [(lo, hi) for lo, hi in DEFAULT_BLOCK_SPANS if lo <= last_col]
    return spans


def split_steps(text):
    """Split the "Hoạt động" field into phases + assessment + differentiation.

    The field holds the whole lesson as one blob of text: numbered phases,
    with "Đánh giá" and "Phân hóa" sections dropped in wherever the author
    happened to write them.
    """
    lines = nfc(str(text or '')).split('\n')

    def is_phase_start(line):
        m = NUMBERED_LINE.match(line)
        if not m:
            return False
        rest = m.group(2).lower()
        return any(kw in rest for kw in PHASE_KEYWORDS)

    def section_label(line):
        return next((name for name, pattern in SECTION_HEADINGS if pattern.match(line)), None)

    sections = {'assessment': [], 'differentiation': []}
    step_lines = []
    current = None
    for line in lines:
        label = section_label(line)
        if label:
            current = label
            sections[label].append(line.strip())
            continue
        if current and is_phase_start(line):
            current = None
        if current:
            sections[current].append(line)
        else:
            step_lines.append(line)

    phases = []
    for line in step_lines:
        if is_phase_start(line) or not phases:
            phases.append([])
        phases[-1].append(line)

    phases = [t for t in ('\n'.join(p).strip() for p in phases) if t]
    return (
        phases,
        '\n'.join(sections['assessment']).strip(),
        '\n'.join(sections['differentiation']).strip(),
    )


def parse_content(content_path, week=None):
    """Parse the content workbook into {activity name: activity dict}.

    When `week` is given, only that week's section is read.
    """
    wb = openpyxl.load_workbook(content_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    last_col = ws.max_column

    activities = {}
    current_week = None
    current_category = None

    for row in range(1, ws.max_row + 1):
        label = label_key(ws.cell(row=row, column=1).value)
        if not label:
            continue

        if is_section_header(ws, row, last_col):
            week_match = WEEK_HEADER.match(label)
            if week_match:
                current_week = int(week_match.group(1))
                current_category = None
            else:
                current_category = label
            continue

        if label not in HEADER_ROW_LABELS:
            continue
        if week is not None and current_week != week:
            continue

        # This is the field-label row: titles sit above it, values below it.
        title_row, data_row = row - 1, row + 1
        for min_col, max_col in block_spans(ws, title_row, last_col):
            activity = parse_block(ws, title_row, row, data_row, min_col, max_col)
            if not activity:
                continue
            activity['category'] = current_category
            activity['week'] = current_week
            activities.setdefault(activity['name'], activity)

    return activities


def parse_block(ws, title_row, header_row, data_row, min_col, max_col):
    """Parse one activity block. Returns None when the block is empty."""
    title = nfc(str(ws.cell(row=title_row, column=min_col).value or '')).strip()
    title_match = ACTIVITY_TITLE.match(title)
    name = (title_match.group(1) if title_match else title).strip().strip('"“”')
    if not name:
        return None

    fields = {}
    for col in range(min_col, max_col + 1):
        key = FIELD_LABELS.get(label_key(ws.cell(row=header_row, column=col).value))
        value = ws.cell(row=data_row, column=col).value
        if not key or value is None:
            continue
        value = nfc(str(value)).strip()
        if value and key not in fields:
            fields[key] = value

    if not fields.get('steps'):
        return None

    phases, assessment, differentiation = split_steps(fields['steps'])
    return {
        'name': name,
        'teacher': fields.get('teacher', ''),
        'standard': fields.get('standard', ''),
        'objectives': fields.get('objectives', ''),
        'criteria': fields.get('criteria', ''),
        'materials': fields.get('materials') or fields.get('prep', ''),
        'phases': phases,
        'assessment': assessment,
        'differentiation': differentiation,
    }


# ============================================================
# SCHEDULE PARSER
# ============================================================

def clean_type(value):
    """Normalize a schedule activity type.

    A type cell can carry a qualifier on a second line or in brackets, e.g.
    "Hoạt động tập thể\\n  Vivokids/thể chất" — only the first line names
    the activity category.
    """
    first_line = nfc(str(value or '')).strip().split('\n')[0]
    return re.sub(r'\s*\(.*?\)\s*', ' ', first_line).strip()


def should_skip(act_type, detail):
    """Skip foreign-teacher slots, free play, routines and unknown types."""
    combined = nfc(f'{act_type} {detail}')
    if any(nfc(kw) in combined for kw in SKIP_KEYWORDS):
        return True
    return label_key(act_type) not in TYPE_TO_CATEGORY


def parse_schedule(schedule_path, class_name):
    """Read one class timetable: metadata plus the activities of each day."""
    wb = openpyxl.load_workbook(schedule_path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.strip() == class_name.strip()), None)
    if not sheet_name:
        available = ', '.join(get_all_classes(schedule_path))
        print(f"Lớp '{class_name}' không tìm thấy. Các lớp có sẵn: {available}")
        sys.exit(1)
    ws = wb[sheet_name]

    week_line = nfc(str(ws.cell(row=2, column=2).value or ''))
    date_match = re.search(r'(\d+)/(\d+)\s*đến\s*(\d+)/(\d+)/(\d+)', week_line)
    if not date_match:
        print(f'Không thể phân tích ngày từ: {week_line}')
        sys.exit(1)
    start_date = datetime(int(date_match.group(5)), int(date_match.group(2)), int(date_match.group(1)))

    theme_raw = nfc(str(ws.cell(row=3, column=2).value or ''))
    theme_match = re.match(r'Chủ đề/Theme:\s*(.+?)(?:\s*\(|/|$)', theme_raw)
    theme = theme_match.group(1).strip() if theme_match else theme_raw.strip()

    teacher_raw = nfc(str(ws.cell(row=3, column=7).value or ''))
    teacher_match = re.search(r'Giáo viên:\s*(.+)', teacher_raw)
    teacher = teacher_match.group(1).split('-')[0].strip().rstrip("' ") if teacher_match else ''

    days = []
    for day_idx in range(5):
        col = DAY_COLUMNS[day_idx]
        activities = []
        for type_row, detail_row in TIME_SLOTS:
            act_type = clean_type(ws.cell(row=type_row, column=col).value)
            detail = nfc(str(ws.cell(row=detail_row, column=col).value or '')).strip()
            if act_type and not should_skip(act_type, detail):
                activities.append({'type': act_type, 'detail': detail})

        days.append({
            'day_num': day_idx + 2,  # Thứ 2 = Monday
            'date': start_date + timedelta(days=day_idx),
            'activities': activities,
        })

    return {'class_name': class_name, 'theme': theme, 'teacher': teacher, 'days': days}


# ============================================================
# ACTIVITY MATCHING
# ============================================================

def candidate_names(detail):
    """Names to try when looking up a schedule entry in the content."""
    detail = nfc(detail)
    names = [detail, normalize_name(detail)]
    for part in detail.split('\n'):
        part = part.strip()
        if part:
            names.extend([part, normalize_name(part)])
    return [n for n in names if n]


def find_activity(schedule_entry, activities):
    """Find the content of a schedule entry, preferring its own category."""
    category = TYPE_TO_CATEGORY.get(label_key(schedule_entry['type']))
    pools = [{k: v for k, v in activities.items() if v['category'] == category}, activities]

    for pool in pools:
        if not pool:
            continue
        for name in candidate_names(schedule_entry['detail']):
            for stored, activity in pool.items():
                if nfc(normalize_name(name)).lower() == nfc(normalize_name(stored)).lower():
                    return activity

        best, best_score = None, 0
        for name in candidate_names(schedule_entry['detail']):
            for stored, activity in pool.items():
                score = match_score(normalize_name(name), normalize_name(stored))
                if score > best_score and score >= 60:
                    best, best_score = activity, score
        if best:
            return best

    return None


# ============================================================
# SHEET BUILDER
# ============================================================

def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    """Apply styles to a cell."""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def style_range(ws, min_row, max_row, min_col, max_col, font=None, fill=None, alignment=None, border=None):
    """Apply styles to a rectangular range of cells."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            apply_style(ws.cell(row=row, column=col), font, fill, alignment, border)


def summary_name(detail):
    """Activity name as shown in the summary table.

    A physical-education slot names both tracks in one cell:
    "TC&SK/Vivokids và vui chơi / TC& SK: <name> / Vivokids: FUNNY GAME".
    Only the TC&SK line names the lesson being planned.
    """
    if 'TC&SK' in detail or 'Vivokids' in detail:
        for line in detail.split('\n'):
            line = line.strip()
            if line.startswith(('TC& SK:', 'TC&SK:')):
                return line.split(':', 1)[1].strip()
    return detail


def write_summary_table(ws, entries, start_row):
    """Write the summary table of the day. Returns the next row."""
    r = start_row

    for col, header in enumerate(['STT', 'Loại hoạt động', 'Tên hoạt động', 'Mục tiêu hoạt động'], start=1):
        if col == 4:
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
        apply_style(ws.cell(row=r, column=col, value=header),
                    FONT_BOLD, alignment=ALIGN_CENTER, border=BORDER_MEDIUM)
    style_range(ws, r, r, 5, 9, border=BORDER_MEDIUM)
    r += 1

    for idx, entry in enumerate(entries, start=1):
        activity = entry.get('content')
        apply_style(ws.cell(row=r, column=1, value=idx),
                    FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_MEDIUM)
        apply_style(ws.cell(row=r, column=2, value=entry['type']),
                    FONT_NORMAL, alignment=ALIGN_LEFT_CENTER, border=BORDER_MEDIUM)
        apply_style(ws.cell(row=r, column=3, value=summary_name(entry['detail'])),
                    FONT_NORMAL, alignment=ALIGN_LEFT_CENTER, border=BORDER_MEDIUM)

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
        apply_style(ws.cell(row=r, column=4, value=activity['objectives'] if activity else ''),
                    FONT_NORMAL, alignment=ALIGN_LEFT_CENTER, border=BORDER_MEDIUM)
        style_range(ws, r, r, 5, 9, border=BORDER_MEDIUM)
        r += 1

    return r


def write_detail_header(ws, start_row):
    """Write the two header rows of the detail table. Returns the next row."""
    r = start_row
    headers = {
        1: 'HOẠT ĐỘNG', 2: 'CHUẨN/MỤC TIÊU CHƯƠNG',
        3: 'MỤC TIÊU BÀI/ MỤC TIÊU HOẠT ĐỘNG', 4: 'TIÊU CHÍ THÀNH CÔNG',
        5: 'TÀI LIỆU HỌC TẬP, CHUẨN BỊ', 6: 'CÁC BƯỚC TỔ CHỨC HOẠT ĐỘNG',
        8: 'ĐÁNH GIÁ', 9: 'PHÂN HÓA',
    }

    # Every column but F/G spans both header rows; F:G is split into
    # "Hoạt động" and "Mục tiêu hướng tới" on the second row.
    for col in [1, 2, 3, 4, 5, 8, 9]:
        ws.merge_cells(start_row=r, start_column=col, end_row=r + 1, end_column=col)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

    for col, text in headers.items():
        apply_style(ws.cell(row=r, column=col, value=text),
                    FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER, BORDER_MEDIUM)
    apply_style(ws.cell(row=r, column=7), FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER, BORDER_MEDIUM)
    r += 1

    open_bottom = Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE)
    apply_style(ws.cell(row=r, column=6, value='Hoạt động'),
                FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER, open_bottom)
    apply_style(ws.cell(row=r, column=7, value='Mục tiêu hướng tới'),
                FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER, open_bottom)
    for col in [1, 2, 3, 4, 5, 8, 9]:
        apply_style(ws.cell(row=r, column=col), FONT_WHITE_BOLD, FILL_NAVY, ALIGN_CENTER, BORDER_MEDIUM)

    return r + 1


def objective_targets(objectives):
    """Build the "Mục tiêu hướng tới" value from the lesson objectives."""
    count = len([line for line in objectives.split('\n') if line.strip()])
    if count <= 1:
        return 'MT 1'
    return ', '.join(f'MT{i}' for i in range(1, min(count, 4) + 1))


def write_activity_block(ws, entry, start_row):
    """Write the detail block of one activity. Returns the next row."""
    activity = entry.get('content')
    act_type = CATEGORY_TITLES.get(activity['category'], entry['type']) if activity else entry['type']

    if not activity:
        name = next((p.strip() for p in reversed(entry['detail'].split('\n')) if p.strip()), '')
        ws.cell(row=start_row, column=1, value=f'{act_type}\n{name}'.strip())
        for col in range(1, 10):
            apply_style(ws.cell(row=start_row, column=col), FONT_NORMAL,
                        alignment=ALIGN_LEFT_TOP, border=BORDER_THIN)
        return start_row + 1

    phases = activity['phases'] or ['']
    end_row = start_row + len(phases) - 1

    values = {
        1: f"{act_type}\n{activity['name']}",
        2: activity['standard'],
        3: activity['objectives'],
        4: activity['criteria'],
        5: activity['materials'],
        7: objective_targets(activity['objectives']),
        8: activity['assessment'],
        9: activity['differentiation'],
    }
    for col, value in values.items():
        ws.cell(row=start_row, column=col, value=value)
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

    for offset, phase in enumerate(phases):
        ws.cell(row=start_row + offset, column=6, value=phase)

    for row in range(start_row, end_row + 1):
        for col in range(1, 10):
            apply_style(ws.cell(row=row, column=col), FONT_NORMAL,
                        alignment=ALIGN_LEFT_TOP, border=BORDER_THIN)

    return end_row + 1


def write_header(ws, schedule, day, location, start_row=1):
    """Write the header block (rows 1-7). Returns the next row."""
    date = day['date']
    lines = [
        ('GIÁO ÁN', FONT_TITLE, None, ALIGN_CENTER),
        (f'Dành cho thứ {day["day_num"]}, ngày {date.strftime("%d/%m/%Y")}',
         FONT_GRAY_BOLD, None, ALIGN_CENTER),
        (schedule['theme'], FONT_THEME, FILL_RED, ALIGN_CENTER),
        (f'Cơ sở: {location}', FONT_BOLD, None, ALIGN_LEFT_CENTER),
        (f'Giáo viên thực hiện: {schedule["teacher"]}', FONT_BOLD, None, ALIGN_LEFT_CENTER),
        (f'Lớp: {schedule["class_name"]}', FONT_BOLD, None, ALIGN_LEFT_CENTER),
        ('Các hoạt động trong ngày:', FONT_BOLD, FILL_GREEN, ALIGN_CENTER),
    ]

    r = start_row
    for text, font, fill, alignment in lines:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        apply_style(ws.cell(row=r, column=1, value=text), font, fill, alignment, BORDER_MEDIUM)
        style_range(ws, r, r, 2, 9, border=BORDER_MEDIUM)
        r += 1

    ws.row_dimensions[start_row].height = 48
    return r


def write_footer(ws, start_row):
    """Write the teacher reflection footer."""
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
    apply_style(ws.cell(row=start_row, column=1, value='Suy ngẫm về dạy & học:'),
                FONT_FOOTER, alignment=Alignment(wrap_text=True), border=BORDER_MEDIUM)
    style_range(ws, start_row, start_row, 2, 9, border=BORDER_MEDIUM)
    ws.row_dimensions[start_row].height = 63
    return start_row + 1


def set_sheet_properties(ws):
    """Column widths and page setup of an output sheet."""
    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width
    ws.sheet_format.defaultColWidth = 14.4
    ws.sheet_format.defaultRowHeight = 15.0
    ws.page_setup.orientation = 'landscape'
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


def build_sheet(ws, schedule, day, entries, location):
    """Build one weekday sheet: header, summary table, detail table, footer."""
    r = write_header(ws, schedule, day, location)
    r = write_summary_table(ws, entries, r)
    r = write_detail_header(ws, r)
    for entry in entries:
        r = write_activity_block(ws, entry, r)
    write_footer(ws, r)
    set_sheet_properties(ws)


# ============================================================
# MAIN
# ============================================================

def get_all_classes(schedule_path):
    """Class sheets of a schedule file.

    A schedule workbook also carries unrelated sheets (evidence matrices,
    review notes); only sheets with a week line in B2 are class timetables.
    """
    wb = openpyxl.load_workbook(schedule_path, data_only=True)
    classes = []
    for sheet in wb.sheetnames:
        week_line = nfc(str(wb[sheet].cell(row=2, column=2).value or ''))
        if re.search(r'\d+/\d+\s*đến\s*\d+/\d+/\d+', week_line):
            classes.append(sheet.strip())
    return classes


def list_classes(schedule_path):
    """Print the class sheets of a schedule file."""
    print(f"Các lớp có trong '{schedule_path}':")
    for name in get_all_classes(schedule_path):
        print(f'  - {name}')


def week_of(schedule_path, class_name):
    """Week number from the theme line of a schedule sheet, e.g. "(Tuần 1)"."""
    wb = openpyxl.load_workbook(schedule_path, data_only=True)
    for sheet in wb.sheetnames:
        if sheet.strip() != class_name.strip():
            continue
        theme = nfc(str(wb[sheet].cell(row=3, column=2).value or ''))
        match = re.search(r'\(\s*tuần\s*(\d+)\s*\)', theme, re.IGNORECASE)
        if match:
            return int(match.group(1))
    return None


def generate(schedule_path, content_path, class_names, output_dir, location, week=None):
    """Generate one lesson plan file per class."""
    os.makedirs(output_dir, exist_ok=True)

    for class_name in class_names:
        print(f"\n{'=' * 60}")
        print(f"Lớp '{class_name}'")
        schedule = parse_schedule(schedule_path, class_name)

        class_week = week if week is not None else week_of(schedule_path, class_name)
        activities = parse_content(content_path, class_week)
        week_label = f'tuần {class_week}' if class_week else 'tất cả các tuần'
        print(f"  Nội dung: {len(activities)} hoạt động ({week_label})")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for day in schedule['days']:
            date = day['date']
            sheet_name = f'Thứ {day["day_num"]}_{date.strftime("%d.%m.%Y")}'

            entries = []
            for entry in day['activities']:
                content = find_activity(entry, activities)
                entries.append({**entry, 'content': content})
                status = '✓' if content else '✗ (không có nội dung)'
                detail = entry['detail'].replace('\n', ' ')[:40]
                print(f"  {sheet_name}: {entry['type']} - {detail}... {status}")

            if not entries:
                print(f'  {sheet_name}: Không có hoạt động nào')
                continue

            build_sheet(wb.create_sheet(title=sheet_name), schedule, day, entries, location)

        output_path = os.path.join(output_dir, f"giao_an_{class_name.replace(' ', '_')}.xlsx")
        wb.save(output_path)
        print(f'Đã tạo: {output_path}')


def main():
    parser = argparse.ArgumentParser(
        description='Giáo Án Generator - Tạo giáo án từ kế hoạch tuần và file "KH chi tiết".',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ví dụ:
  %(prog)s -s "Tuần 1.xlsx" -c "KH chi tiết.xlsx" --class "Vega 6"
  %(prog)s -s "Tuần 1.xlsx" -c "KH chi tiết.xlsx" --all -o output/
  %(prog)s -s "Tuần 1.xlsx" --list-classes
        """,
    )
    parser.add_argument('--schedule', '-s', required=True, metavar='FILE',
                        help='File kế hoạch giảng dạy theo tuần (.xlsx)')
    parser.add_argument('--content', '-c', metavar='FILE',
                        help='File nội dung chi tiết "KH chi tiết" (.xlsx)')

    group = parser.add_mutually_exclusive_group()
    group.add_argument('--class', '-C', dest='class_name', metavar='NAME',
                       help='Tên lớp (ví dụ: "Vega 6")')
    group.add_argument('--all', '-a', action='store_true',
                       help='Tạo giáo án cho tất cả các lớp')
    group.add_argument('--list-classes', '-l', action='store_true',
                       help='Liệt kê các lớp có trong file kế hoạch')

    parser.add_argument('--week', '-w', type=int, metavar='N',
                        help='Số tuần trong file nội dung (mặc định: lấy từ chủ đề của lớp)')
    parser.add_argument('--location', default=DEFAULT_LOCATION, metavar='TEXT',
                        help='Tên cơ sở ghi trên giáo án')
    parser.add_argument('--output', '-o', default='output', metavar='DIR',
                        help='Thư mục đầu ra (mặc định: output/)')

    args = parser.parse_args()

    if not os.path.isfile(args.schedule):
        parser.error(f"Không tìm thấy file: '{args.schedule}'")

    if args.list_classes:
        list_classes(args.schedule)
        return

    if not args.content and not args.class_name and not args.all:
        list_classes(args.schedule)
        print('\nSử dụng --class hoặc --all cùng với --content để tạo giáo án.')
        return

    if not args.content:
        parser.error('Cần cung cấp --content (-c) khi tạo giáo án')
    if not os.path.isfile(args.content):
        parser.error(f"Không tìm thấy file: '{args.content}'")

    class_names = get_all_classes(args.schedule) if args.all else [args.class_name]
    generate(args.schedule, args.content, class_names, args.output, args.location, args.week)


if __name__ == '__main__':
    main()
